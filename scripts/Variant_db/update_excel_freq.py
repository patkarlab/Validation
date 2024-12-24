#!/usr/bin/env python3

from openpyxl import load_workbook
import os
import sys
import csv

input_excel = sys.argv[1]	# input excel file
freq_file = sys.argv[2]		# tsv file with chrm:pos:ref:alt and frequency

wb_col = load_workbook(input_excel)
sheet = wb_col['append_final_concat']
map_dict = dict()
freq_col = "db_freq"

def key_gen (sheet_name, row):
	chrom = sheet_name.cell(row = row, column=1).value
	pos = sheet_name.cell(row = row, column=2).value
	ref = sheet_name.cell(row = row, column=4).value
	alt = sheet_name.cell(row = row, column=5).value
	key = chrom + ":" + str(pos) + ":" + ref + ":" + alt
	return key

# Check the last column of the sheet, add a new column named db_freq if not already present
last_column_name = sheet.cell(row = 1, column=sheet.max_column).value
if freq_col not in last_column_name:
	sheet.cell(row = 1, column = sheet.max_column + 1, value = freq_col)

for row in range(2, sheet.max_row + 1):
	key = key_gen(sheet, row) 
	map_dict[key] = -1	# Make a dict with default value of -1

# Read the freq file and update the values
with open (freq_file, 'r') as freq:
	freq_handle = csv.reader(freq, delimiter="\t")
	for lines in freq_handle:
		key_val = lines[0]
		freq_val = lines[1]
		if key_val in map_dict:
			map_dict[key_val] = freq_val	# Update the dict with freq values
			
for row in range(2, sheet.max_row + 1):
	key2 = key_gen(sheet, row)
	sheet.cell(row = row, column = sheet.max_column, value = float(map_dict[key2]))	

wb_col.save(input_excel)
